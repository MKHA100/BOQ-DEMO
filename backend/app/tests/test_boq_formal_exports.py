from __future__ import annotations

import fitz
from openpyxl import load_workbook

from app.tests.test_boq import setup_boq_data


def test_professional_pdf_and_excel_exports(foundation_db):
    from app.boq.service import boq_service
    from app.boq.repo import boq_repository
    from app.jobs.worker import process_one
    from app.storage.storage_service import storage_service

    project, *_ = setup_boq_data()
    boq_service.refresh(project["id"])
    pdf_request = boq_service.request_export(project["id"], {"format": "pdf", "floor_mode": "combined", "floor_id": None}, None)
    xlsx_request = boq_service.request_export(project["id"], {"format": "xlsx", "floor_mode": "combined", "floor_id": None}, None)
    process_one("pdf-export-test", ["export.generate"])
    process_one("xlsx-export-test", ["export.generate"])

    pdf = boq_repository.get_export(project["id"], pdf_request["export"]["id"])
    xlsx = boq_repository.get_export(project["id"], xlsx_request["export"]["id"])
    assert pdf["status"] == "ready" and xlsx["status"] == "ready"

    pdf_path = storage_service.key_to_path(pdf["object_key"])
    with fitz.open(pdf_path) as document:
        text = "\n".join(page.get_text() for page in document)
        assert "Main Summary" in text
        assert "Source Traceability" in text

    workbook = load_workbook(storage_service.key_to_path(xlsx["object_key"]), data_only=False)
    assert {"MAIN SUMMARY", "PRELIMINARIES", "ELEMENT PROPERTIES", "SOURCE TRACEABILITY", "NEEDS REVIEW"}.issubset(set(workbook.sheetnames))
    assert any(name.startswith("BILL ") for name in workbook.sheetnames)


def test_floor_breakdown_exports_include_floor_sections(foundation_db):
    from app.boq.service import boq_service
    from app.boq.repo import boq_repository
    from app.jobs.worker import process_one
    from app.storage.storage_service import storage_service

    project, *_ = setup_boq_data()
    boq_service.refresh(project["id"], grouping_mode="floor")
    pdf_request = boq_service.request_export(project["id"], {"format": "pdf", "floor_mode": "floor_breakdown", "floor_id": None}, None)
    xlsx_request = boq_service.request_export(project["id"], {"format": "xlsx", "floor_mode": "floor_breakdown", "floor_id": None}, None)
    process_one("floor-pdf-export-test", ["export.generate"])
    process_one("floor-xlsx-export-test", ["export.generate"])

    pdf = boq_repository.get_export(project["id"], pdf_request["export"]["id"])
    xlsx = boq_repository.get_export(project["id"], xlsx_request["export"]["id"])
    with fitz.open(storage_service.key_to_path(pdf["object_key"])) as document:
        text = "\n".join(page.get_text() for page in document)
        assert "Floor Breakdown" in text
    workbook = load_workbook(storage_service.key_to_path(xlsx["object_key"]), data_only=False)
    assert any(name.startswith("FLOOR - ") for name in workbook.sheetnames)
