from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

THIN = Side(style="thin", color="94A3B8")
MEDIUM = Side(style="medium", color="475569")
HEADER_FILL = PatternFill("solid", fgColor="DCE6F1")
SECTION_FILL = PatternFill("solid", fgColor="E2E8F0")
REVIEW_FILL = PatternFill("solid", fgColor="FEF3C7")


def _safe_sheet_name(value: str, used: set[str]) -> str:
    base = re.sub(r"[\\/*?:\[\]]", "-", value).strip()[:31] or "Bill"
    name = base; index = 2
    while name in used:
        suffix = f" {index}"; name = f"{base[:31-len(suffix)]}{suffix}"; index += 1
    used.add(name); return name


def _title(sheet, report: dict, title: str, columns: int) -> int:
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=columns)
    cell = sheet.cell(1, 1, title); cell.font = Font(name="Arial", size=15, bold=True); cell.alignment = Alignment(horizontal="center")
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=columns)
    cell = sheet.cell(2, 1, report.get("project_name") or "Project"); cell.font = Font(name="Arial", size=11, bold=True); cell.alignment = Alignment(horizontal="center")
    sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=columns)
    cell = sheet.cell(3, 1, f"Client: {report.get('client_name') or '—'}    Location: {report.get('location') or '—'}"); cell.alignment = Alignment(horizontal="center")
    return 5


def _style_header(sheet, row: int, columns: int) -> None:
    for col in range(1, columns + 1):
        cell = sheet.cell(row, col); cell.font = Font(bold=True); cell.fill = HEADER_FILL
        cell.border = Border(top=MEDIUM, bottom=MEDIUM, left=THIN, right=THIN); cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_table(sheet, start_row: int, end_row: int, columns: int) -> None:
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=columns):
        for cell in row:
            cell.border = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _print_settings(sheet, repeat_row: int | None = None, landscape: bool = True) -> None:
    sheet.freeze_panes = f"A{(repeat_row or 5) + 1}"
    sheet.sheet_view.showGridLines = False
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.orientation = "landscape" if landscape else "portrait"
    sheet.page_setup.fitToWidth = 1; sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.oddFooter.center.text = "Page &P of &N"
    if repeat_row: sheet.print_title_rows = f"{repeat_row}:{repeat_row}"



def _floor_items(report: dict) -> dict[str, list[dict]]:
    floors: dict[str, list[dict]] = {}
    for bill in report.get("bills") or []:
        for section in bill.get("sections") or []:
            for item in section.get("items") or []:
                for floor in item.get("floor_names") or ["Unassigned"]:
                    floors.setdefault(str(floor), []).append({
                        **item,
                        "bill_no": bill.get("bill_no"),
                        "bill_name": bill.get("name"),
                        "section_code": section.get("code"),
                        "section_name": section.get("name"),
                    })
    return floors


def _write_floor_sheets(workbook, used: set[str], report: dict, currency: str) -> None:
    if report.get("export_floor_mode") != "floor_breakdown":
        return
    include_rates = bool(report.get("include_rates")); include_amounts = bool(report.get("include_amounts"))
    for floor_name, items in sorted(_floor_items(report).items()):
        columns = 6 + int(include_rates) + int(include_amounts)
        sheet = workbook.create_sheet(_safe_sheet_name(f"FLOOR - {floor_name}", used))
        row = _title(sheet, report, f"Floor Breakdown – {floor_name}", columns)
        headers = ["Bill", "Section", "Item", "Description", "Unit", "Quantity"]
        if include_rates: headers.append("Rate")
        if include_amounts: headers.append("Amount")
        sheet.append(headers); _style_header(sheet, row, columns)
        for item in items:
            data = [item.get("bill_no"), item.get("section_code"), item.get("item_number"), item.get("description"), item.get("unit"), item.get("quantity")]
            if include_rates: data.append(item.get("rate"))
            if include_amounts: data.append(item.get("amount"))
            sheet.append(data)
            if item.get("status") == "needs_review":
                for cell in sheet[sheet.max_row]: cell.fill = REVIEW_FILL
        if include_amounts:
            sheet.append(["", "", "", "Floor total", "", ""] + ([""] if include_rates else []) + [sum(float(item.get("amount") or 0) for item in items)])
            for cell in sheet[sheet.max_row]: cell.font = Font(bold=True); cell.border = Border(top=MEDIUM)
        _style_table(sheet, row + 1, sheet.max_row, columns)
        widths = [10, 14, 14, 70, 10, 14, 18, 20]
        for index in range(1, columns + 1): sheet.column_dimensions[get_column_letter(index)].width = widths[index - 1]
        for current in range(row + 1, sheet.max_row + 1):
            sheet.cell(current, 6).number_format = "0.000"
            if include_rates: sheet.cell(current, 7).number_format = f'"{currency}" #,##0.00'
            if include_amounts: sheet.cell(current, columns).number_format = f'"{currency}" #,##0.00'
        _print_settings(sheet, row)

def write_xlsx(path: Path, report: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook(); workbook.remove(workbook.active); used: set[str] = set()
    currency = str(report.get("currency") or "Rs")

    summary = workbook.create_sheet(_safe_sheet_name("MAIN SUMMARY", used))
    row = _title(summary, report, report.get("title") or "Bill of Quantities", 4)
    summary.append(["Bill", "Description", "Items", "Amount"]); _style_header(summary, row, 4)
    for bill in report.get("bills") or []:
        summary.append([bill.get("bill_no"), bill.get("name"), bill.get("item_count"), bill.get("subtotal") if report.get("include_amounts") else None])
    values = report.get("summary") or {}
    summary.append(["", "Subtotal", "", values.get("subtotal") if report.get("include_amounts") else None])
    summary.append(["", f"VAT ({float(report.get('vat_percentage') or 0):g}%)", "", values.get("vat") if report.get("include_amounts") else None])
    summary.append(["", "Grand Total", "", values.get("grand_total") if report.get("include_amounts") else None])
    _style_table(summary, row + 1, summary.max_row, 4)
    summary.column_dimensions["A"].width = 12; summary.column_dimensions["B"].width = 55; summary.column_dimensions["C"].width = 12; summary.column_dimensions["D"].width = 20
    for cell in summary[summary.max_row]: cell.font = Font(bold=True); cell.border = Border(top=MEDIUM, bottom=MEDIUM)
    for r in range(row + 1, summary.max_row + 1): summary.cell(r, 4).number_format = f'"{currency}" #,##0.00'
    _print_settings(summary, row)

    if report.get("include_preliminaries"):
        sheet = workbook.create_sheet(_safe_sheet_name("PRELIMINARIES", used)); row = _title(sheet, report, "Preliminaries", 3)
        sheet.append(["No.", "Notes", ""]); _style_header(sheet, row, 3)
        for index, note in enumerate(report.get("preliminaries", {}).get("notes") or [], start=1): sheet.append([index, note, ""])
        row = sheet.max_row + 2; sheet.cell(row, 1, "Mode of payment").font = Font(bold=True, size=11)
        row += 1; sheet.append(["Category", "Description", ""]); _style_header(sheet, row, 3)
        for item in report.get("preliminaries", {}).get("mode_of_payment") or []: sheet.append([item.get("category"), item.get("description"), ""])
        _style_table(sheet, 6, sheet.max_row, 3); sheet.column_dimensions["A"].width = 14; sheet.column_dimensions["B"].width = 100; sheet.column_dimensions["C"].width = 4
        _print_settings(sheet, 5)

    for bill in report.get("bills") or []:
        sheet = workbook.create_sheet(_safe_sheet_name(f"BILL {bill.get('bill_no')}", used))
        include_rates = bool(report.get("include_rates")); include_amounts = bool(report.get("include_amounts"))
        columns = 4 + int(include_rates) + int(include_amounts)
        row = _title(sheet, report, f"Bill {bill.get('bill_no')} – {bill.get('name')}", columns)
        headers = ["Item", "Description", "Unit", "Quantity"]
        if include_rates: headers.append("Rate")
        if include_amounts: headers.append("Amount")
        sheet.append(headers); _style_header(sheet, row, columns)
        for section in bill.get("sections") or []:
            section_row = sheet.max_row + 1
            sheet.merge_cells(start_row=section_row, start_column=1, end_row=section_row, end_column=columns)
            cell = sheet.cell(section_row, 1, f"{section.get('code')} – {section.get('name')}"); cell.font = Font(bold=True); cell.fill = SECTION_FILL
            for item in section.get("items") or []:
                data = [item.get("item_number"), item.get("description"), item.get("unit"), item.get("quantity")]
                if include_rates: data.append(item.get("rate"))
                if include_amounts: data.append(item.get("amount"))
                sheet.append(data)
                current = sheet.max_row
                if item.get("status") == "needs_review":
                    for cell in sheet[current]: cell.fill = REVIEW_FILL
            if include_amounts:
                data = ["", "Section subtotal", "", ""]
                if include_rates: data.append("")
                data.append(section.get("subtotal")); sheet.append(data)
                for cell in sheet[sheet.max_row]: cell.font = Font(bold=True)
        _style_table(sheet, row + 1, sheet.max_row, columns)
        widths = [14, 78, 10, 14, 18, 20]
        for index in range(1, columns + 1): sheet.column_dimensions[get_column_letter(index)].width = widths[index - 1]
        for r in range(row + 1, sheet.max_row + 1):
            sheet.cell(r, 4).number_format = "0.000"
            if include_rates: sheet.cell(r, 5).number_format = f'"{currency}" #,##0.00'
            if include_amounts: sheet.cell(r, columns).number_format = f'"{currency}" #,##0.00'
        _print_settings(sheet, row)

    _write_floor_sheets(workbook, used, report, currency)

    properties = workbook.create_sheet(_safe_sheet_name("ELEMENT PROPERTIES", used)); row = _title(properties, report, "Element Properties", 8)
    headers = ["Source item", "Type", "Type code", "Floor", "Width mm", "Height mm", "Material", "Finish"]
    properties.append(headers); _style_header(properties, row, len(headers))
    for item in report.get("element_properties") or []:
        properties.append([item.get("source_item"), item.get("element_type"), item.get("type_code"), item.get("floor"), item.get("width_mm"), item.get("height_mm"), item.get("material"), item.get("finish")])
    _style_table(properties, row + 1, properties.max_row, len(headers));
    for i, width in enumerate([16, 14, 14, 20, 12, 12, 28, 28], start=1): properties.column_dimensions[get_column_letter(i)].width = width
    _print_settings(properties, row)

    trace = workbook.create_sheet(_safe_sheet_name("SOURCE TRACEABILITY", used)); row = _title(trace, report, "Source Traceability", 6)
    headers = ["BOQ item", "Description", "Source items", "Floors", "Status", "Missing fields"]
    trace.append(headers); _style_header(trace, row, 6)
    for item in report.get("source_traceability") or []:
        trace.append([item.get("boq_item_number"), item.get("description"), ", ".join(item.get("source_items") or []), ", ".join(item.get("floors") or []), item.get("status"), ", ".join(item.get("missing_fields") or [])])
    _style_table(trace, row + 1, trace.max_row, 6)
    for i, width in enumerate([14, 70, 30, 22, 16, 35], start=1): trace.column_dimensions[get_column_letter(i)].width = width
    _print_settings(trace, row)

    review = workbook.create_sheet(_safe_sheet_name("NEEDS REVIEW", used)); row = _title(review, report, "Needs Review", 4)
    headers = ["BOQ item", "Description", "Missing fields", "Source items"]
    review.append(headers); _style_header(review, row, 4)
    for item in report.get("needs_review") or []:
        review.append([item.get("boq_item_number"), item.get("description"), ", ".join(item.get("missing_fields") or []), ", ".join(item.get("source_items") or [])])
        for cell in review[review.max_row]: cell.fill = REVIEW_FILL
    _style_table(review, row + 1, review.max_row, 4)
    for i, width in enumerate([14, 75, 35, 30], start=1): review.column_dimensions[get_column_letter(i)].width = width
    _print_settings(review, row)

    workbook.save(path)
